from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="003366")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ROW_FILL_LIGHT = PatternFill("solid", fgColor="DCE6F1")
ROW_FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
NUM_FORMAT = '#,##0'


def apply_header(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_row(ws, row_idx, values, num_cols=None):
    fill = ROW_FILL_LIGHT if row_idx % 2 == 0 else ROW_FILL_WHITE
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
    return fill


def set_num_format(ws, row_idx, col_indices):
    for col in col_indices:
        cell = ws.cell(row=row_idx, column=col)
        if cell.value is not None and cell.value != "":
            cell.number_format = NUM_FORMAT


def auto_column_width(ws, min_width=12):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2


# ---------------------------------------------------------------------------
# HOJA 1 — Conciliacion
# ---------------------------------------------------------------------------
CONCILIACION_HEADERS = [
    "fecha_pago", "monto_ars", "pais_pos", "reserva", "cliente",
    "precio_usd", "ars_estimado", "diferencia", "porcentaje", "nota",
]

CONCILIACION_DATA = [
    ("03/01/2026", 63000,  "Australia",   5672943398, "Yui Hashimoto",            50.40,  74088,  -11088, "-15%",   "Match confirmado manual"),
    ("03/01/2026", 73000,  "Japan",        5692907081, "Deniz Nedret Karagülle",   42.72,  62798,  10202,  "+16%",   "Match confirmado manual"),
    ("07/01/2026", 12900,  "Israel",       5703087091, "Yuval Teper",              8.90,   13083,  -183,   "-1.4%",  "Match automatico"),
    ("08/01/2026", 45000,  "Israel",       5701543421, "Sophie Belozovsky",        32.04,  47099,  -2099,  "-4.5%",  "Match automatico"),
    ("10/01/2026", 47000,  "Israel",       5701543421, "Sophie Belozovsky",        32.04,  47099,  -99,    "-0.2%",  "Revisar posible doble cobro"),
    ("10/01/2026", 72800,  "Australia",   6288499925, "Anna Petrie",              49.56,  72853,  -53,    "-0.1%",  "Match automatico"),
    ("11/01/2026", 13000,  "China",        5798587974, "Miao Ma",                  9.00,   13230,  -230,   "-1.7%",  "Match automatico"),
    ("16/01/2026", 26700,  "Japan",        5764247919, "Tadayuki Kiyohara",        17.70,  26019,  681,    "+2.6%",  "Match automatico"),
    ("17/01/2026", 117800, "Israel",       5506717743, "ליאם נגר",                 None,   None,   None,   None,     "Extension no registrada en Booking"),
    ("17/01/2026", 113000, "Israel",       5506717743, "Acompañante de ליאם נגר",  None,   None,   None,   None,     "Extension no registrada en Booking"),
    ("24/01/2026", 25500,  "Spain",        5447963935, "Angela Monteiro",          17.80,  26166,  -666,   "-2.5%",  "Match automatico"),
    ("24/01/2026", 89000,  "France",       5447938358, "Tristan Gaudicheau",       62.23,  91478,  -2478,  "-2.7%",  "Match automatico"),
    ("26/01/2026", 25800,  "South Korea",  5821951576, "Hyunmuk Park",             18.00,  26460,  -660,   "-2.5%",  "Match automatico"),
    ("26/01/2026", 30600,  "France",       6146048631, "Salomé Guémon",            17.80,  26166,  4434,   "+17%",   "Match confirmado manual"),
    ("26/01/2026", 29300,  "UK",           6312944882, "Marie Egan",               21.36,  31399,  -2099,  "-6.7%",  "Match automatico"),
    ("28/01/2026", 12900,  "South Korea",  5565574234, "OH SANGYOON",              8.90,   13083,  -183,   "-1.4%",  "Match automatico"),
    ("28/01/2026", 13000,  "Israel",       6016837187, "ram amit",                 26.70,  39249,  None,   None,     "Pago 1/3 huespedes"),
    ("28/01/2026", 13000,  "Israel",       6016837187, "ram amit",                 26.70,  39249,  None,   None,     "Pago 2/3 huespedes"),
    ("28/01/2026", 13000,  "Israel",       6016837187, "ram amit",                 26.70,  39249,  -249,   "-0.6%",  "Pago 3/3 - total 39000"),
    ("28/01/2026", 13500,  "Israel",       5798536690, "Maya Brown",               8.90,   13083,  417,    "+3.2%",  "Match automatico"),
    ("29/01/2026", 41800,  "Spain",        5410050860, "Isidoro Fernandez Magaña", 29.04,  42689,  -889,   "-2.1%",  "Match automatico"),
    ("31/01/2026", 14000,  "France",       5505123347, "Emerick Bon",              10.68,  15700,  -1700,  "-10.8%", "Match automatico limite"),
    ("01/02/2026", 93000,  "France",       6702517178, "Raphael Garcia",           64.56,  91998,  1002,   "+1.1%",  "Match automatico"),
    ("02/02/2026", 12600,  "Israel",       6928004320, "Isidoro Fernandez Magaña", 9.68,   13794,  -1194,  "-8.7%",  "Match automatico"),
    ("02/02/2026", 27000,  "Japan",        5239599768, "Michiya Higuchi",          17.90,  25508,  1492,   "+5.9%",  "Match automatico"),
    ("03/02/2026", 12900,  "Israel",       6175673614, "Itay Mosafi",              8.90,   12683,  217,    "+1.7%",  "Match automatico"),
    ("03/02/2026", 14700,  "Israel",       6849318058, "ram amit",                 19.58,  27901,  None,   None,     "Pago 1/2"),
    ("03/02/2026", 14700,  "Israel",       6849318058, "ram amit",                 19.58,  27901,  1499,   "+5.4%",  "Pago 2/2 - total 29400"),
    ("04/02/2026", 46400,  "France",       6790359786, "Thomas Bourret",           32.04,  45657,  743,    "+1.6%",  "Match automatico"),
    ("05/02/2026", 90000,  "USA",          6702772758, "Roy Eliav",                53.40,  76095,  13905,  "+18%",   "Match confirmado manual"),
    ("06/02/2026", 12600,  "Israel",       6647067228, "Yuval Beham Eran",         10.80,  15390,  -2790,  "-18%",   "Match confirmado manual"),
    ("07/02/2026", 16300,  "Israel",       6365205984, "Juan I. Petroselli",       10.00,  14250,  2050,   "+14%",   "Match confirmado manual"),
    ("08/02/2026", 12900,  "France",       5894973646, "Victor Gradea",            9.00,   12825,  75,     "+0.6%",  "Match automatico"),
    ("11/02/2026", 15000,  "France",       5034869111, "Clemence Fabrizi",         10.80,  15390,  -390,   "-2.5%",  "Match automatico"),
    ("12/02/2026", 30000,  "France",       6084569150, "Julie Martin",             21.60,  30780,  -780,   "-2.5%",  "Match automatico"),
    ("12/02/2026", 37000,  "France",       6245194816, "Gaelle Berguin",           26.70,  38048,  -1048,  "-2.8%",  "Match automatico"),
    ("17/02/2026", 121000, "Brazil",       6192569781, "Gonzalo Mercado",          88.40,  129948, -8948,  "-6.9%",  "Match automatico"),
]

# ---------------------------------------------------------------------------
# HOJA 2 — Pagos sin reserva
# ---------------------------------------------------------------------------
PAGOS_SIN_RESERVA_HEADERS = ["fecha_pago", "monto_ars", "pais", "observacion"]

PAGOS_SIN_RESERVA_DATA = [
    ("05/01/2026", 23000, "Israel",      "Sin reserva identificable"),
    ("10/01/2026", 11000, "Argentina",   "Walk-in local"),
    ("13/01/2026", 64000, "Israel",      "Diferencia demasiado grande (-32%)"),
    ("14/01/2026", 16000, "Guatemala",   "Sin reserva identificable"),
    ("19/01/2026", 16900, "Argentina",   "Walk-in local"),
    ("21/01/2026", 25000, "South Korea", "Sin match (+93% sobre Lee Seojin)"),
    ("22/01/2026", 14700, "Italy",       "Sin match (-37% de marongiu)"),
    ("22/01/2026", 14700, "Italy",       "Sin match (idem)"),
    ("25/01/2026", 25500, "Spain",       "Posible duplicado del 24/01 - revisar"),
    ("26/01/2026", 10000, "Israel",      "Monto bajo, posible seña"),
    ("26/01/2026", 28000, "Belgium",     "Sin reserva belga identificada"),
    ("27/01/2026", 15300, "France",      "Guemon ya asignada - sin match"),
    ("27/01/2026", 35000, "USA",         "Sin match (-26% de Chris Acosta)"),
    ("28/01/2026", 25500, "Spain",       "Fernandez Magaña ya asignado al 29/01"),
    ("02/02/2026", 66000, "Brazil",      "Pago parcial previo de Gonzalo Mercado (ver Hoja 4)"),
    ("02/02/2026", 11500, "Japan",       "Higuchi ya asignado - sin match"),
    ("03/02/2026", 13900, "Spain",       "Sin reserva española en esa fecha"),
    ("03/02/2026", 8000,  "France",      "Monto bajo, posible diferencia"),
    ("04/02/2026", 88000, "France",      "Bourret ya asignado - posible doble cobro"),
    ("04/02/2026", 29400, "Israel",      "ram amit ya asignado - posible doble cobro"),
    ("04/02/2026", 26000, "Australia",   "Sin reserva australiana en febrero"),
    ("09/02/2026", 2000,  "Israel",      "Seña o diferencia"),
    ("09/02/2026", 13300, "Israel",      "Sin match"),
    ("09/02/2026", 12600, "USA",         "Sin match"),
    ("09/02/2026", 8400,  "Israel",      "Sin match"),
    ("10/02/2026", 12600, "Israel",      "Roy Eliav ya asignado - sin match"),
    ("11/02/2026", 54000, "SIN TARJETA", "Dinero en cuenta - no es cobro POS"),
]

# ---------------------------------------------------------------------------
# HOJA 3 — Reservas sin pago
# ---------------------------------------------------------------------------
RESERVAS_SIN_PAGO_HEADERS = [
    "reserva", "cliente", "entrada", "precio_usd", "ars_estimado", "observacion",
]

RESERVAS_SIN_PAGO_DATA = [
    (5506717743, "ליאם נגר",               "14/01/2026", 63.84, 93845,  "Noches originales cobradas fuera del sistema. Extension registrada en Hoja 1 (17/01)"),
    (6074538550, "antonio marongiu",       "17/01/2026", 31.68, 46570,  "Pagos 14700x2 no coinciden - posible cobro online"),
    (5853361028, "Lee Seojin",             "20/01/2026", 8.80,  12936,  "Pago 25000 del 21/01 no coincide - posible cobro online"),
    (5167331493, "xabi legras",            "23/01/2026", 53.40, 78498,  "Sin pago POS identificado"),
    (6811984007, "Mehdi Benechebli",       "23/01/2026", 21.36, 31399,  "Pago 28000 fuera de rango - posible cobro online"),
    (6880283247, "cyprien Benoit",         "26/01/2026", 21.36, 31399,  "Sin pago POS identificado"),
    (6932328395, "Chris Acosta",           "26/01/2026", 32.04, 47099,  "Pago 35000 fuera de rango - posible cobro online"),
    (6234537085, "Shira Shaked Ben David", "28/01/2026", 10.68, 15700,  "Sin pago POS identificado"),
    (6192569781, "Gonzalo Mercado",        "29/01/2026", 88.40, 129948, "Pagos 66000 (02/02) + 121000 (17/02) - ver Hoja 4"),
    (6219547182, "kaori hashinaga",        "02/02/2026", 10.80, 15390,  "Sin pago POS identificado"),
    (6702783183, "Marie Du Fayet",         "04/02/2026", 18.00, 25650,  "Sin pago POS identificado"),
    (5078659986, "YAMIL LUCERO",           "10/02/2026", 10.80, 15390,  "Sin pago POS identificado"),
    (5012047553, "Guido del Amo",          "11/02/2026", 32.04, 45657,  "Pago Dinero en cuenta 54000 - no es POS"),
    (6116330950, "STEFFANY SANTIAGO",      "10/02/2026", 17.80, 25350,  "Pago 13300 no coincide"),
    (6069573290, "Daiana Ceballos",        "14/02/2026", 27.00, 38475,  "Sin pago POS en el periodo"),
]

# ---------------------------------------------------------------------------
# HOJA 4 — Pagos multiples
# ---------------------------------------------------------------------------
PAGOS_MULTIPLES_HEADERS = [
    "reserva", "cliente", "precio_usd", "ars_estimado",
    "pagos_pos", "total_pagado", "diferencia", "tipo",
]

PAGOS_MULTIPLES_DATA = [
    (6016837187, "ram amit (enero)",   26.70, 39249,  "13000 + 13000 + 13000", 39000,  -249,  "3 huespedes misma reserva"),
    (6849318058, "ram amit (febrero)", 19.58, 27901,  "14700 + 14700",         29400,  1499,  "2 pagos - 2 camas"),
    (5506717743, "ליאם נגר",           63.84, 93845,  "117800 + 113000",       230800, None,  "Extension no registrada en Booking"),
    (5701543421, "Sophie Belozovsky",  32.04, 47099,  "45000 + 47000",         92000,  44901, "Posible doble cobro - REVISAR"),
    (6790359786, "Thomas Bourret",     32.04, 45657,  "46400 + 88000",         134400, 88743, "Posible doble cobro - REVISAR"),
    (6192569781, "Gonzalo Mercado",    88.40, 129948, "66000 + 121000",        187000, 57052, "Posible extension o pago por dos personas"),
]


def build_sheet(ws, title, headers, data, num_col_indices):
    ws.title = title
    ws.row_dimensions[1].height = 20
    apply_header(ws, headers)
    for i, row_values in enumerate(data, 2):
        apply_row(ws, i, row_values)
        set_num_format(ws, i, num_col_indices)
    auto_column_width(ws)


def main():
    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    # monto_ars=col2, ars_estimado=col7, diferencia=col8
    build_sheet(ws1, "Conciliacion", CONCILIACION_HEADERS, CONCILIACION_DATA, [2, 7, 8])

    # Sheet 2
    ws2 = wb.create_sheet()
    # monto_ars=col2
    build_sheet(ws2, "Pagos sin reserva", PAGOS_SIN_RESERVA_HEADERS, PAGOS_SIN_RESERVA_DATA, [2])

    # Sheet 3
    ws3 = wb.create_sheet()
    # ars_estimado=col5
    build_sheet(ws3, "Reservas sin pago", RESERVAS_SIN_PAGO_HEADERS, RESERVAS_SIN_PAGO_DATA, [5])

    # Sheet 4
    ws4 = wb.create_sheet()
    # ars_estimado=col4, total_pagado=col6, diferencia=col7
    build_sheet(ws4, "Pagos multiples", PAGOS_MULTIPLES_HEADERS, PAGOS_MULTIPLES_DATA, [4, 6, 7])

    filename = "conciliacion_pagos_booking_enero_febrero_2026.xlsx"
    wb.save(filename)
    print(f"✅ Archivo generado: {filename}")


if __name__ == "__main__":
    main()
